# -*- coding: utf-8 -*-
import os
import csv
import math
import random
import time
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
try:
    from generate_terrain import generate_terrain  
except ImportError:
    print("CRITICAL ERROR: Could not find 'generate_terrain.py'.")
    print("Ensure both files are in the same folder.")
    exit(1)

class Config:
    def __init__(self):
        self.IMG_SIZE = 129
        self.GRID_SIZE = self.IMG_SIZE
        self.REAL_SIZE = 50.0
        self.MAX_HEIGHT = 4.0
        self.ROBOT_MASS = 5.0  # kg 
        self.GRAVITY = 9.81     # m/s^2

        self.HEIGHTMAP_BOTTOM_ALIGNED = False
        self.REMOVE_GROUND_PLANE = True
        self.HOVER_HEIGHT = 0.05
        self.ROAD_THICKNESS = 0.12
        self.ROAD_WIDTH = 0.30
        self.RESAMPLE_STEP_M = 0.30
        self.ADD_DEBUG_PINS = False
        self.PIN_EVERY_N = 8
        self.START = (random.randrange(5, 60), random.randrange(5, 60))
        self.GOAL = (115, 115)
        self.MIN_FRICTION = 0.008 
        self.MAX_SLOPE = 1.0
        self.CURRENT_DIR = os.getcwd()
        self.OUTPUT_TEXTURE = os.path.join(self.CURRENT_DIR, "terrain_texture.png")
        self.OUTPUT_HEIGHT_IMG = os.path.join(self.CURRENT_DIR, "heightmap.png")
        self.OUTPUT_WORLD = os.path.join(self.CURRENT_DIR, "thesis_3d_path.world")

class Terrain:
    def __init__(self, config, height_img, friction_map):
        self.config = config
        self.height_img = height_img
        self.friction_map = friction_map

    def bilinear_sample_u8(self, img: Image.Image, u: float, v: float) -> float:
        W, H = img.size
        x = max(0.0, min(1.0, u)) * (W - 1)
        y = max(0.0, min(1.0, v)) * (H - 1)

        x0 = int(math.floor(x)); x1 = min(x0 + 1, W - 1)
        y0 = int(math.floor(y)); y1 = min(y0 + 1, H - 1)
        dx = x - x0; dy = y - y0

        p00 = img.getpixel((x0, y0)) / 255.0
        p10 = img.getpixel((x1, y0)) / 255.0
        p01 = img.getpixel((x0, y1)) / 255.0
        p11 = img.getpixel((x1, y1)) / 255.0

        p0 = p00*(1-dx) + p10*dx
        p1 = p01*(1-dx) + p11*dx
        return p0*(1-dy) + p1*dy

    def csv_index_to_uv(self, grid_x: int, grid_y: int) -> tuple:
        u = grid_x / (self.config.GRID_SIZE - 1)
        v = grid_y / (self.config.GRID_SIZE - 1)
        return u, v

    def world_xy_from_grid(self, grid_x: int, grid_y: int) -> tuple:
        u, v = self.csv_index_to_uv(grid_x, grid_y)
        world_x = (u - 0.5) * self.config.REAL_SIZE
        world_y = (0.5 - v) * self.config.REAL_SIZE
        return world_x, world_y

    def world_to_uv(self, wx: float, wy: float, pos_x=0.0, pos_y=0.0) -> tuple:
        u = (wx - pos_x) / self.config.REAL_SIZE + 0.5
        v = 0.5 - (wy - pos_y) / self.config.REAL_SIZE
        return u, v

    def world_xy_to_grid(self, wx: float, wy: float) -> tuple:
        u, v = self.world_to_uv(wx, wy, pos_x=0.0, pos_y=0.0)
        gx = int(max(0, min(self.config.GRID_SIZE - 1, round(u * (self.config.GRID_SIZE - 1)))))
        gy = int(max(0, min(self.config.GRID_SIZE - 1, round(v * (self.config.GRID_SIZE - 1)))))
        return gx, gy

    def terrain_z_from_world_xy(self, wx: float, wy: float) -> float:
        pos_z = (self.config.MAX_HEIGHT / 2.0) if self.config.HEIGHTMAP_BOTTOM_ALIGNED else 0.0
        u, v = self.world_to_uv(wx, wy, pos_x=0.0, pos_y=0.0)
        p = self.bilinear_sample_u8(self.height_img, u, v)
        return pos_z + (p * self.config.MAX_HEIGHT)

    def get_height_m_from_img(self, grid_x: int, grid_y: int) -> float:
        u, v = self.csv_index_to_uv(grid_x, grid_y)
        p = self.bilinear_sample_u8(self.height_img, u, v)
        return p * self.config.MAX_HEIGHT

    def get_3d_step_cost(self, curr, nxt):
        cx, cy = curr
        nx, ny = nxt
        dist_2d_grid = math.hypot(nx - cx, ny - cy)
        if dist_2d_grid == 0:
            return float('inf')
        # Convert grid distance to real-world meters
        meters_per_cell = self.config.REAL_SIZE / (self.config.GRID_SIZE - 1)
        dist_m = dist_2d_grid * meters_per_cell

        friction = self.friction_map[nx][ny]
        h_curr = self.get_height_m_from_img(cx, cy)
        h_next = self.get_height_m_from_img(nx, ny)
        dh = h_next - h_curr
        
        # Check physical slope constraint
        slope = abs(dh) / dist_m
        if slope > self.config.MAX_SLOPE:
            return float('inf')
        weight = self.config.ROBOT_MASS * self.config.GRAVITY
        
        # Work done against friction
        energy_friction = friction * weight * dist_m
        energy_gravity = (weight * dh) if dh > 0 else 0.0
        
        return energy_friction + energy_gravity

class AStarPlanner:
    def __init__(self, config, terrain):
        self.config = config
        self.terrain = terrain

    def heuristic(self, curr, goal, start):
        cx, cy = curr
        gx, gy = goal
        sx, sy = start
        
        # Convert 2D grid distance to meters
        meters_per_cell = self.config.REAL_SIZE / (self.config.GRID_SIZE - 1)
        dist_2d_grid = math.hypot(gx - cx, gy - cy)
        dist_m = dist_2d_grid * meters_per_cell
        
        cz = self.terrain.get_height_m_from_img(cx, cy)
        gz = self.terrain.get_height_m_from_img(gx, gy)
        dh = gz - cz
        
        weight = self.config.ROBOT_MASS * self.config.GRAVITY
        
        # Admissible energy components
        min_e_friction = self.config.MIN_FRICTION * weight * dist_m
        min_e_gravity = (weight * dh) if dh > 0 else 0.0
        
        h_joules = min_e_friction + min_e_gravity
        
        #tie-breaker to prevent searching symmetrical paths
        dx1 = cx - gx
        dy1 = cy - gy
        dx2 = sx - gx
        dy2 = sy - gy
        cross_product = abs(dx1 * dy2 - dx2 * dy1)
        
        return h_joules + (cross_product * 0.001)

    def run_astar(self):
        import heapq
        open_list = []
        counter = 0
        heapq.heappush(open_list, (0, counter, self.config.START))

        came_from = {}
        g_score = {self.config.START: 0.0}
        f_score = {self.config.START: self.heuristic(self.config.START, self.config.GOAL, self.config.START)}
        closed = set()

        while open_list:
            current = heapq.heappop(open_list)[2]
            if current == self.config.GOAL:
                path = []
                while current in came_from:
                    path.append(current)
                    current = came_from[current]
                path.append(self.config.START)
                path.reverse()
                return path

            closed.add(current)
            cx, cy = current
            neighbors = [
                (cx, cy-1), (cx, cy+1), (cx-1, cy), (cx+1, cy),
                (cx-1, cy-1), (cx+1, cy-1), (cx-1, cy+1), (cx+1, cy+1)
            ]
            for nx, ny in neighbors:
                if not (0 <= nx < self.config.GRID_SIZE and 0 <= ny < self.config.GRID_SIZE):
                    continue
                if (nx, ny) in closed:
                    continue
                step = self.terrain.get_3d_step_cost(current, (nx, ny))
                if step == float('inf'):
                    continue
                tentative_g = g_score[current] + step
                if tentative_g < g_score.get((nx, ny), float('inf')):
                    came_from[(nx, ny)] = current
                    g_score[(nx, ny)] = tentative_g
                    f_score[(nx, ny)] = tentative_g + self.heuristic((nx, ny), self.config.GOAL, self.config.START)
                    counter += 1
                    heapq.heappush(open_list, (f_score[(nx, ny)], counter, (nx, ny)))

        return []  # no path

class PathProcessor:
    def __init__(self, config, terrain):
        self.config = config
        self.terrain = terrain

    def get_bresenham_line(self, x0, y0, x1, y1):
        points = []
        dx = abs(x1 - x0)
        dy = abs(y1 - y0)
        x, y = x0, y0
        sx = -1 if x0 > x1 else 1
        sy = -1 if y0 > y1 else 1
        if dx > dy:
            err = dx / 2.0
            while x != x1:
                points.append((x, y))
                err -= dy
                if err < 0:
                    y += sy
                    err += dx
                x += sx
        else:
            err = dy / 2.0
            while y != y1:
                points.append((x, y))
                err -= dx
                if err < 0:
                    x += sx
                    err += dy
                y += sy
        points.append((x, y))
        return points

    def get_dense_grid_cost(self, path_sparse):
        total_cost = 0.0
        is_valid = True
        for i in range(len(path_sparse) - 1):
            line = self.get_bresenham_line(path_sparse[i][0], path_sparse[i][1], path_sparse[i+1][0], path_sparse[i+1][1])
            for j in range(len(line) - 1):
                cost = self.terrain.get_3d_step_cost(line[j], line[j+1])
                if cost == float('inf'):
                    is_valid = False
                total_cost += cost
        return total_cost, is_valid

    def smooth_path_los(self, path):
        if len(path) <= 2:
            return path
            
        print("Smoothing path (Cost-Aware String-Pulling)...")
        smoothed_path = [path[0]]
        current_index = 0
        
        while current_index < len(path) - 1:
            furthest_visible = current_index + 1
            
            for j in range(current_index + 2, len(path)):
                shortcut_line = [path[current_index], path[j]]
                straight_cost, valid = self.get_dense_grid_cost(shortcut_line)
                
                astar_segment = path[current_index:j+1]
                astar_cost, _ = self.get_dense_grid_cost(astar_segment)
                
                if valid and straight_cost <= (astar_cost * 1.02):
                    furthest_visible = j
                else:
                    break
                    
            smoothed_path.append(path[furthest_visible])
            current_index = furthest_visible
            
        print(f"Original nodes: {len(path)} -> Smoothed nodes: {len(smoothed_path)}")
        return smoothed_path

    def optimize_path_rubberband(self, path, max_iters=30):
        if len(path) <= 2:
            return path
            
        print("Optimizing path (Rubberband Gradient Descent)...")
        current_path = path.copy()
        optimized = True
        iterations = 0
        
        while optimized and iterations < max_iters:
            optimized = False
            iterations += 1
            
            for i in range(1, len(current_path) - 1):
                prev_node = current_path[i-1]
                curr_node = current_path[i]
                next_node = current_path[i+1]
                
                cost_curr1, v1 = self.get_dense_grid_cost([prev_node, curr_node])
                cost_curr2, v2 = self.get_dense_grid_cost([curr_node, next_node])
                best_cost = cost_curr1 + cost_curr2
                best_node = curr_node
                
                cx, cy = curr_node
                neighbors = [
                    (cx, cy-1), (cx, cy+1), (cx-1, cy), (cx+1, cy),
                    (cx-1, cy-1), (cx+1, cy-1), (cx-1, cy+1), (cx+1, cy+1)
                ]
                
                for nx, ny in neighbors:
                    if not (0 <= nx < self.config.GRID_SIZE and 0 <= ny < self.config.GRID_SIZE): 
                        continue
                        
                    c1, val1 = self.get_dense_grid_cost([prev_node, (nx, ny)])
                    c2, val2 = self.get_dense_grid_cost([(nx, ny), next_node])
                    
                    if val1 and val2 and (c1 + c2) < best_cost - 0.0001:
                        best_cost = c1 + c2
                        best_node = (nx, ny)
                        
                if best_node != curr_node:
                    current_path[i] = best_node
                    optimized = True
                    
        print(f"Rubberband Optimization finished in {iterations} iterations.")
        return current_path

class MetricsCalculator:
    def __init__(self, config, terrain):
        self.config = config
        self.terrain = terrain

    def get_path_metrics(self, path, timing_data):
        total_dist_2d = 0.0
        total_dist_3d = 0.0
        min_z = float('inf')
        max_z = float('-inf')
        max_step_dz = 0.0
        pos_z_for_sdf = (self.config.MAX_HEIGHT / 2.0) if self.config.HEIGHTMAP_BOTTOM_ALIGNED else 0.0
        
        for i in range(len(path)):
            cx, cy = path[i]
            wx, wy = self.terrain.world_xy_from_grid(cx, cy)
            cz = self.terrain.terrain_z_from_world_xy(wx, wy)
            
            if cz < min_z: min_z = cz
            if cz > max_z: max_z = cz
            
            if i > 0:
                px, py = path[i-1]
                pwx, pwy = self.terrain.world_xy_from_grid(px, py)
                pz = self.terrain.terrain_z_from_world_xy(pwx, pwy)
                
                dist2d = math.hypot(wx - pwx, wy - pwy)
                total_dist_2d += dist2d
                
                dz = cz - pz
                total_dist_3d += math.hypot(dist2d, dz)
                
                if abs(dz) > max_step_dz:
                    max_step_dz = abs(dz)
                    
        overall_elev_change = max_z - min_z
        
        processor = PathProcessor(self.config, self.terrain)
        true_cost, is_valid = processor.get_dense_grid_cost(path)
        
        return {
            'astar_time': timing_data['astar'],
            'smooth_time': timing_data['smooth'],
            'opt_time': timing_data['opt'],
            'total_time': timing_data['total'],
            'dist_2d': total_dist_2d,
            'dist_3d': total_dist_3d,
            'energy': true_cost,
            'valid': is_valid,
            'min_z': min_z,
            'max_z': max_z,
            'elev_span': overall_elev_change,
            'max_step_dz': max_step_dz
        }

class StandardAStarPlanner:
    def __init__(self, config, terrain):
        self.config = config
        self.terrain = terrain

    def standard_cost(self, curr, nxt):
        cx, cy = curr; nx, ny = nxt
        dist_2d = math.hypot(nx - cx, ny - cy)
        if dist_2d == 0: return float('inf')
        h_curr = self.terrain.get_height_m_from_img(cx, cy)
        h_next = self.terrain.get_height_m_from_img(nx, ny)
        dh = h_next - h_curr
        if abs(dh) / dist_2d > self.config.MAX_SLOPE: return float('inf')
        return math.sqrt(dist_2d**2 + dh**2)

    def standard_heuristic_func(self, curr, goal):
        cx, cy = curr; gx, gy = goal
        cz = self.terrain.get_height_m_from_img(cx, cy)
        gz = self.terrain.get_height_m_from_img(gx, gy)
        return math.sqrt((gx - cx)**2 + (gy - cy)**2 + (gz - cz)**2)

    def run_standard_astar(self):
        import heapq
        open_list = []
        counter = 0
        heapq.heappush(open_list, (0, counter, self.config.START))
        came_from = {}
        g_score = {self.config.START: 0.0}
        f_score = {self.config.START: self.standard_heuristic_func(self.config.START, self.config.GOAL)}
        closed = set()

        while open_list:
            current = heapq.heappop(open_list)[2]
            if current == self.config.GOAL:
                path = []
                while current in came_from:
                    path.append(current)
                    current = came_from[current]
                path.append(self.config.START)
                path.reverse()
                return path

            closed.add(current)
            cx, cy = current
            neighbors = [(cx, cy-1), (cx, cy+1), (cx-1, cy), (cx+1, cy), (cx-1, cy-1), (cx+1, cy-1), (cx-1, cy+1), (cx+1, cy+1)]
            for nx, ny in neighbors:
                if not (0 <= nx < self.config.GRID_SIZE and 0 <= ny < self.config.GRID_SIZE): continue
                if (nx, ny) in closed: continue
                step = self.standard_cost(current, (nx, ny))
                if step == float('inf'): continue
                tentative_g = g_score[current] + step
                if tentative_g < g_score.get((nx, ny), float('inf')):
                    came_from[(nx, ny)] = current
                    g_score[(nx, ny)] = tentative_g
                    f_score[(nx, ny)] = tentative_g + self.standard_heuristic_func((nx, ny), self.config.GOAL)
                    counter += 1
                    heapq.heappush(open_list, (f_score[(nx, ny)], counter, (nx, ny)))
        return []

class DijkstraPlanner:
    def __init__(self, config, terrain):
        self.config = config
        self.terrain = terrain

    def standard_cost(self, curr, nxt):
        cx, cy = curr; nx, ny = nxt
        dist_2d = math.hypot(nx - cx, ny - cy)
        if dist_2d == 0: return float('inf')
        h_curr = self.terrain.get_height_m_from_img(cx, cy)
        h_next = self.terrain.get_height_m_from_img(nx, ny)
        dh = h_next - h_curr
        if abs(dh) / dist_2d > self.config.MAX_SLOPE: return float('inf')
        return math.sqrt(dist_2d**2 + dh**2)

    def run_dijkstra(self):
        import heapq
        open_list = []
        counter = 0
        heapq.heappush(open_list, (0, counter, self.config.START))
        came_from = {}
        g_score = {self.config.START: 0.0}
        f_score = {self.config.START: 0.0} 
        closed = set()

        while open_list:
            current = heapq.heappop(open_list)[2]
            if current == self.config.GOAL:
                path = []
                while current in came_from:
                    path.append(current)
                    current = came_from[current]
                path.append(self.config.START)
                path.reverse()
                return path

            closed.add(current)
            cx, cy = current
            neighbors = [(cx, cy-1), (cx, cy+1), (cx-1, cy), (cx+1, cy), (cx-1, cy-1), (cx+1, cy-1), (cx-1, cy+1), (cx+1, cy+1)]
            for nx, ny in neighbors:
                if not (0 <= nx < self.config.GRID_SIZE and 0 <= ny < self.config.GRID_SIZE): continue
                if (nx, ny) in closed: continue
                step = self.standard_cost(current, (nx, ny))
                if step == float('inf'): continue
                tentative_g = g_score[current] + step
                if tentative_g < g_score.get((nx, ny), float('inf')):
                    came_from[(nx, ny)] = current
                    g_score[(nx, ny)] = tentative_g
                    f_score[(nx, ny)] = tentative_g 
                    counter += 1
                    heapq.heappush(open_list, (f_score[(nx, ny)], counter, (nx, ny)))
        return []

def main():
    map_type = input("Enter scenario type (mix/friction/hill): ").strip().lower()

    # Generate terrain once
    friction_map, height_map, height_img = generate_terrain(map_type)
    
    results = []
    
    for run in range(10):
        print(f"Run {run + 1}/10")
        config = Config()  # New random start each run
        terrain = Terrain(config, height_img, friction_map)
        planner = AStarPlanner(config, terrain)
        processor = PathProcessor(config, terrain)
        metrics = MetricsCalculator(config, terrain)
        
        # Custom A*
        t0 = time.time()
        raw_path = planner.run_astar()
        t1 = time.time()
        if not raw_path:
            print(f"Custom A* no path for run {run + 1}")
            continue
        smoothed_path = processor.smooth_path_los(raw_path)
        t2 = time.time()
        final_path = processor.optimize_path_rubberband(smoothed_path)
        t3 = time.time()
        
        timing_data = {
            'astar': t1 - t0,
            'smooth': t2 - t1,
            'opt': t3 - t2,
            'total': t3 - t0
        }
        custom_metrics = metrics.get_path_metrics(final_path, timing_data)
        custom_metrics['run'] = run + 1
        custom_metrics['algorithm'] = 'Custom A*'
        results.append(custom_metrics)
        
        # Standard A*
        standard_planner = StandardAStarPlanner(config, terrain)
        t_n0 = time.time()
        raw_normal_path = standard_planner.run_standard_astar()
        t_n1 = time.time()
        if not raw_normal_path:
            print(f"Standard A* no path for run {run + 1}")
            continue
        smooth_normal_path = processor.smooth_path_los(raw_normal_path)
        t_n2 = time.time()
        normal_path = processor.optimize_path_rubberband(smooth_normal_path)
        t_n3 = time.time()
        
        timing_data_n = {
            'astar': t_n1 - t_n0,
            'smooth': t_n2 - t_n1,
            'opt': t_n3 - t_n2,
            'total': t_n3 - t_n0
        }
        normal_metrics = metrics.get_path_metrics(normal_path, timing_data_n)
        normal_metrics['run'] = run + 1
        normal_metrics['algorithm'] = 'Standard A*'
        results.append(normal_metrics)
        
        # Dijkstra
        dijkstra_planner = DijkstraPlanner(config, terrain)
        t_d0 = time.time()
        raw_dijkstra_path = dijkstra_planner.run_dijkstra()
        t_d1 = time.time()
        if not raw_dijkstra_path:
            print(f"Dijkstra no path for run {run + 1}")
            continue
        smooth_dijkstra_path = processor.smooth_path_los(raw_dijkstra_path)
        t_d2 = time.time()
        dijkstra_path = processor.optimize_path_rubberband(smooth_dijkstra_path)
        t_d3 = time.time()
        
        timing_data_d = {
            'astar': t_d1 - t_d0,
            'smooth': t_d2 - t_d1,
            'opt': t_d3 - t_d2,
            'total': t_d3 - t_d0
        }
        dijkstra_metrics = metrics.get_path_metrics(dijkstra_path, timing_data_d)
        dijkstra_metrics['run'] = run + 1
        dijkstra_metrics['algorithm'] = 'Dijkstra'
        results.append(dijkstra_metrics)
    
    # Write to Excel
    if results:
        fieldnames = ['run', 'algorithm', 'astar_time', 'smooth_time', 'opt_time', 'total_time', 'dist_2d', 'dist_3d', 'energy', 'valid', 'min_z', 'max_z', 'elev_span', 'max_step_dz']
        excel_filename = f'experiment_results_{map_type}.xlsx'
        write_to_excel(excel_filename, fieldnames, results)
        print(f"Results saved to {excel_filename}")
        compute_averages_excel(excel_filename)
        compute_comparative_stats_excel(excel_filename)
        generate_plots(excel_filename, map_type)
    else:
        print("No results to save.")

def write_to_excel(filename, fieldnames, results):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        start_row = 1
        for col_idx, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col_idx, value=field)
        start_row = 2
    
    for row_idx, result in enumerate(results, start_row):
        for col_idx, field in enumerate(fieldnames, 1):
            ws.cell(row=row_idx, column=col_idx, value=result.get(field, ''))
    
    wb.save(filename)

def compute_averages_excel(excel_file):
    from collections import defaultdict
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    data = defaultdict(list)
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    
    for row in range(2, ws.max_row + 1):
        algo = ws.cell(row, headers.get('algorithm')).value
        if algo:
            for key in ['astar_time', 'smooth_time', 'opt_time', 'total_time', 'dist_2d', 'dist_3d', 'energy', 'min_z', 'max_z', 'elev_span', 'max_step_dz']:
                if key in headers:
                    val = ws.cell(row, headers[key]).value
                    if val is not None:
                        data[algo + '_' + key].append(float(val))
            if 'valid' in headers:
                val = ws.cell(row, headers['valid']).value
                data[algo + '_valid'].append(1 if str(val).lower() == 'true' else 0)
    
    print("\n--- AVERAGE METRICS PER ALGORITHM ---")
    algorithms = ['Custom A*', 'Standard A*', 'Dijkstra']
    for algo in algorithms:
        print(f"\n{algo}:")
        for key in ['astar_time', 'smooth_time', 'opt_time', 'total_time', 'dist_2d', 'dist_3d', 'energy', 'min_z', 'max_z', 'elev_span', 'max_step_dz']:
            values = data[algo + '_' + key]
            avg = sum(values) / len(values) if values else 0
            print(f"  {key}: {avg:.4f}")
        valid_values = data[algo + '_valid']
        avg_valid = sum(valid_values) / len(valid_values) if valid_values else 0
        print(f"  valid_fraction: {avg_valid:.4f}")

def compute_comparative_stats_excel(excel_file):
    from collections import defaultdict
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    run_data = defaultdict(dict)
    
    for row in range(2, ws.max_row + 1):
        run_num = ws.cell(row, headers.get('run')).value
        algo = ws.cell(row, headers.get('algorithm')).value
        if run_num and algo:
            run_data[run_num][algo] = {
                'energy': float(ws.cell(row, headers.get('energy')).value or 0),
                'total_time': float(ws.cell(row, headers.get('total_time')).value or 0),
                'dist_2d': float(ws.cell(row, headers.get('dist_2d')).value or 0),
                'dist_3d': float(ws.cell(row, headers.get('dist_3d')).value or 0),
                'valid': str(ws.cell(row, headers.get('valid')).value).lower() == 'true'
            }
    
    diffs_energy_vs_standard = []
    diffs_energy_vs_dijkstra = []
    diffs_time_vs_standard = []
    diffs_time_vs_dijkstra = []
    diffs_dist2d_vs_standard = []
    diffs_dist2d_vs_dijkstra = []
    diffs_dist3d_vs_standard = []
    diffs_dist3d_vs_dijkstra = []
    
    wins_energy_vs_standard = 0
    wins_energy_vs_dijkstra = 0
    wins_time_vs_standard = 0
    wins_time_vs_dijkstra = 0
    wins_dist2d_vs_standard = 0
    wins_dist2d_vs_dijkstra = 0
    wins_dist3d_vs_standard = 0
    wins_dist3d_vs_dijkstra = 0
    
    total_runs = 0
    
    for run, algos in run_data.items():
        if 'Custom A*' not in algos or 'Standard A*' not in algos or 'Dijkstra' not in algos:
            continue
        total_runs += 1
        custom = algos['Custom A*']
        standard = algos['Standard A*']
        dijkstra = algos['Dijkstra']
        
        diff_e_s = custom['energy'] - standard['energy']
        diff_e_d = custom['energy'] - dijkstra['energy']
        diffs_energy_vs_standard.append(diff_e_s)
        diffs_energy_vs_dijkstra.append(diff_e_d)
        if diff_e_s < 0: wins_energy_vs_standard += 1
        if diff_e_d < 0: wins_energy_vs_dijkstra += 1
        
        diff_t_s = custom['total_time'] - standard['total_time']
        diff_t_d = custom['total_time'] - dijkstra['total_time']
        diffs_time_vs_standard.append(diff_t_s)
        diffs_time_vs_dijkstra.append(diff_t_d)
        if diff_t_s < 0: wins_time_vs_standard += 1
        if diff_t_d < 0: wins_time_vs_dijkstra += 1
        
        diff_d2_s = custom['dist_2d'] - standard['dist_2d']
        diff_d2_d = custom['dist_2d'] - dijkstra['dist_2d']
        diffs_dist2d_vs_standard.append(diff_d2_s)
        diffs_dist2d_vs_dijkstra.append(diff_d2_d)
        if diff_d2_s < 0: wins_dist2d_vs_standard += 1
        if diff_d2_d < 0: wins_dist2d_vs_dijkstra += 1
        
        diff_d3_s = custom['dist_3d'] - standard['dist_3d']
        diff_d3_d = custom['dist_3d'] - dijkstra['dist_3d']
        diffs_dist3d_vs_standard.append(diff_d3_s)
        diffs_dist3d_vs_dijkstra.append(diff_d3_d)
        if diff_d3_s < 0: wins_dist3d_vs_standard += 1
        if diff_d3_d < 0: wins_dist3d_vs_dijkstra += 1
    
    print("\n--- COMPARATIVE STATISTICS (Custom A* vs Others) ---")
    print(f"Total comparable runs: {total_runs}")
    
    def print_stats(label, diffs_s, diffs_d, wins_s, wins_d):
        avg_s = sum(diffs_s) / len(diffs_s) if diffs_s else 0
        avg_d = sum(diffs_d) / len(diffs_d) if diffs_d else 0
        win_pct_s = (wins_s / total_runs) * 100 if total_runs else 0
        win_pct_d = (wins_d / total_runs) * 100 if total_runs else 0
        print(f"\n{label}:")
        print(f"  Avg diff vs Standard A*: {avg_s:.4f} ({win_pct_s:.1f}% wins)")
        print(f"  Avg diff vs Dijkstra: {avg_d:.4f} ({win_pct_d:.1f}% wins)")
    
    print_stats("Energy Consumption (Joules, negative = savings)", diffs_energy_vs_standard, diffs_energy_vs_dijkstra, wins_energy_vs_standard, wins_energy_vs_dijkstra)
    print_stats("Total Time (seconds, negative = faster)", diffs_time_vs_standard, diffs_time_vs_dijkstra, wins_time_vs_standard, wins_time_vs_dijkstra)
    print_stats("2D Distance (meters, negative = shorter)", diffs_dist2d_vs_standard, diffs_dist2d_vs_dijkstra, wins_dist2d_vs_standard, wins_dist2d_vs_dijkstra)
    print_stats("3D Distance (meters, negative = shorter)", diffs_dist3d_vs_standard, diffs_dist3d_vs_dijkstra, wins_dist3d_vs_standard, wins_dist3d_vs_dijkstra)
    
    print("\n--- PER-RUN ENERGY CONSUMPTION (Joules) ---")
    print("Run | Custom A* | Standard A* | Dijkstra | Custom vs Std | Custom vs Dijk")
    print("-" * 75)
    for run in sorted(run_data.keys()):
        algos = run_data[run]
        if 'Custom A*' in algos and 'Standard A*' in algos and 'Dijkstra' in algos:
            c_e = algos['Custom A*']['energy']
            s_e = algos['Standard A*']['energy']
            d_e = algos['Dijkstra']['energy']
            diff_s = c_e - s_e
            diff_d = c_e - d_e
            print(f"{run:3d} | {c_e:10.2f} | {s_e:12.2f} | {d_e:8.2f} | {diff_s:+10.2f} | {diff_d:+10.2f}")

def generate_plots(excel_file, map_type):
    from collections import defaultdict
    
    # Create graphs folder structure
    graphs_dir = os.path.join('graphs', map_type)
    os.makedirs(graphs_dir, exist_ok=True)
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    run_data = defaultdict(dict)
    
    for row in range(2, ws.max_row + 1):
        run_num = ws.cell(row, headers.get('run')).value
        algo = ws.cell(row, headers.get('algorithm')).value
        if run_num and algo:
            run_data[run_num][algo] = {
                'energy': float(ws.cell(row, headers.get('energy')).value or 0),
                'total_time': float(ws.cell(row, headers.get('total_time')).value or 0),
                'dist_2d': float(ws.cell(row, headers.get('dist_2d')).value or 0),
                'dist_3d': float(ws.cell(row, headers.get('dist_3d')).value or 0),
            }
    
    # Organize data by algorithm
    algo_data = defaultdict(lambda: {'energy': [], 'time': [], 'dist_2d': [], 'dist_3d': [], 'runs': []})
    for run in sorted(run_data.keys()):
        algos = run_data[run]
        for algo in ['Custom A*', 'Standard A*', 'Dijkstra']:
            if algo in algos:
                algo_data[algo]['energy'].append(algos[algo]['energy'])
                algo_data[algo]['time'].append(algos[algo]['total_time'])
                algo_data[algo]['dist_2d'].append(algos[algo]['dist_2d'])
                algo_data[algo]['dist_3d'].append(algos[algo]['dist_3d'])
                algo_data[algo]['runs'].append(run)
    
    # Plot 1: Energy consumption box plot
    fig1 = plt.figure(figsize=(10, 6))
    energy_data = [algo_data['Custom A*']['energy'], algo_data['Standard A*']['energy'], algo_data['Dijkstra']['energy']]
    bp = plt.boxplot(energy_data, labels=['Custom A*', 'Standard A*', 'Dijkstra'], patch_artist=True)
    for patch, color in zip(bp['boxes'], ['#2ecc71', '#3498db', '#e74c3c']):
        patch.set_facecolor(color)
    plt.ylabel('Energy (Joules)', fontweight='bold', fontsize=12)
    plt.title(f'Energy Consumption Distribution - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.grid(axis='y', alpha=0.3)
    plot1_path = os.path.join(graphs_dir, '01_energy_distribution.png')
    plt.savefig(plot1_path, dpi=300, bbox_inches='tight')
    plt.close(fig1)
    print(f"  Saved: {plot1_path}")
    
    # Plot 2: Energy consumption over runs
    fig2 = plt.figure(figsize=(12, 6))
    plt.plot(algo_data['Custom A*']['runs'], algo_data['Custom A*']['energy'], 'o-', label='Custom A*', linewidth=2, markersize=6, color='#2ecc71')
    plt.plot(algo_data['Standard A*']['runs'], algo_data['Standard A*']['energy'], 's-', label='Standard A*', linewidth=2, markersize=6, color='#3498db')
    plt.plot(algo_data['Dijkstra']['runs'], algo_data['Dijkstra']['energy'], '^-', label='Dijkstra', linewidth=2, markersize=6, color='#e74c3c')
    plt.xlabel('Run Number', fontweight='bold', fontsize=12)
    plt.ylabel('Energy (Joules)', fontweight='bold', fontsize=12)
    plt.title(f'Energy Consumption Over Runs - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.legend(loc='best', fontsize=11)
    plt.grid(alpha=0.3)
    plot2_path = os.path.join(graphs_dir, '02_energy_over_runs.png')
    plt.savefig(plot2_path, dpi=300, bbox_inches='tight')
    plt.close(fig2)
    print(f"  Saved: {plot2_path}")
    
    # Plot 3: Average energy comparison
    fig3 = plt.figure(figsize=(10, 6))
    custom_avg_e = sum(algo_data['Custom A*']['energy']) / len(algo_data['Custom A*']['energy']) if algo_data['Custom A*']['energy'] else 0
    std_avg_e = sum(algo_data['Standard A*']['energy']) / len(algo_data['Standard A*']['energy']) if algo_data['Standard A*']['energy'] else 0
    dij_avg_e = sum(algo_data['Dijkstra']['energy']) / len(algo_data['Dijkstra']['energy']) if algo_data['Dijkstra']['energy'] else 0
    bars = plt.bar(['Custom A*', 'Standard A*', 'Dijkstra'], [custom_avg_e, std_avg_e, dij_avg_e], color=['#2ecc71', '#3498db', '#e74c3c'], alpha=0.8, edgecolor='black', linewidth=1.5)
    plt.ylabel('Average Energy (Joules)', fontweight='bold', fontsize=12)
    plt.title(f'Average Energy Consumption - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.grid(axis='y', alpha=0.3)
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.1f}', ha='center', va='bottom', fontweight='bold', fontsize=11)
    plot3_path = os.path.join(graphs_dir, '03_average_energy.png')
    plt.savefig(plot3_path, dpi=300, bbox_inches='tight')
    plt.close(fig3)
    print(f"  Saved: {plot3_path}")
    
    # Plot 4: Average planning time comparison
    fig4 = plt.figure(figsize=(10, 6))
    custom_avg_t = sum(algo_data['Custom A*']['time']) / len(algo_data['Custom A*']['time']) if algo_data['Custom A*']['time'] else 0
    std_avg_t = sum(algo_data['Standard A*']['time']) / len(algo_data['Standard A*']['time']) if algo_data['Standard A*']['time'] else 0
    dij_avg_t = sum(algo_data['Dijkstra']['time']) / len(algo_data['Dijkstra']['time']) if algo_data['Dijkstra']['time'] else 0
    bars = plt.bar(['Custom A*', 'Standard A*', 'Dijkstra'], [custom_avg_t, std_avg_t, dij_avg_t], color=['#2ecc71', '#3498db', '#e74c3c'], alpha=0.8, edgecolor='black', linewidth=1.5)
    plt.ylabel('Average Time (seconds)', fontweight='bold', fontsize=12)
    plt.title(f'Average Planning Time - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.grid(axis='y', alpha=0.3)
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.4f}', ha='center', va='bottom', fontweight='bold', fontsize=10)
    plot4_path = os.path.join(graphs_dir, '04_average_time.png')
    plt.savefig(plot4_path, dpi=300, bbox_inches='tight')
    plt.close(fig4)
    print(f"  Saved: {plot4_path}")
    
    # Plot 5: Average 2D distance comparison
    fig5 = plt.figure(figsize=(10, 6))
    custom_avg_d2 = sum(algo_data['Custom A*']['dist_2d']) / len(algo_data['Custom A*']['dist_2d']) if algo_data['Custom A*']['dist_2d'] else 0
    std_avg_d2 = sum(algo_data['Standard A*']['dist_2d']) / len(algo_data['Standard A*']['dist_2d']) if algo_data['Standard A*']['dist_2d'] else 0
    dij_avg_d2 = sum(algo_data['Dijkstra']['dist_2d']) / len(algo_data['Dijkstra']['dist_2d']) if algo_data['Dijkstra']['dist_2d'] else 0
    bars = plt.bar(['Custom A*', 'Standard A*', 'Dijkstra'], [custom_avg_d2, std_avg_d2, dij_avg_d2], color=['#2ecc71', '#3498db', '#e74c3c'], alpha=0.8, edgecolor='black', linewidth=1.5)
    plt.ylabel('Average Distance (meters)', fontweight='bold', fontsize=12)
    plt.title(f'Average 2D Path Distance - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.grid(axis='y', alpha=0.3)
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.1f}', ha='center', va='bottom', fontweight='bold', fontsize=11)
    plot5_path = os.path.join(graphs_dir, '05_average_2d_distance.png')
    plt.savefig(plot5_path, dpi=300, bbox_inches='tight')
    plt.close(fig5)
    print(f"  Saved: {plot5_path}")
    
    # Plot 6: Average 3D distance comparison
    fig6 = plt.figure(figsize=(10, 6))
    custom_avg_d3 = sum(algo_data['Custom A*']['dist_3d']) / len(algo_data['Custom A*']['dist_3d']) if algo_data['Custom A*']['dist_3d'] else 0
    std_avg_d3 = sum(algo_data['Standard A*']['dist_3d']) / len(algo_data['Standard A*']['dist_3d']) if algo_data['Standard A*']['dist_3d'] else 0
    dij_avg_d3 = sum(algo_data['Dijkstra']['dist_3d']) / len(algo_data['Dijkstra']['dist_3d']) if algo_data['Dijkstra']['dist_3d'] else 0
    bars = plt.bar(['Custom A*', 'Standard A*', 'Dijkstra'], [custom_avg_d3, std_avg_d3, dij_avg_d3], color=['#2ecc71', '#3498db', '#e74c3c'], alpha=0.8, edgecolor='black', linewidth=1.5)
    plt.ylabel('Average Distance (meters)', fontweight='bold', fontsize=12)
    plt.title(f'Average 3D Path Distance - {map_type.capitalize()} Map', fontweight='bold', fontsize=14)
    plt.grid(axis='y', alpha=0.3)
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.1f}', ha='center', va='bottom', fontweight='bold', fontsize=11)
    plot6_path = os.path.join(graphs_dir, '06_average_3d_distance.png')
    plt.savefig(plot6_path, dpi=300, bbox_inches='tight')
    plt.close(fig6)
    print(f"  Saved: {plot6_path}")
    
    print(f"\nAll plots saved to: {graphs_dir}")

if __name__ == "__main__":
    main()